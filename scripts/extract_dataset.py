"""
从原始 Excel 数据提取样本, 保存为 NPZ。

输入: data/LTai1.xlsx ~ LTai5.xlsx
  每个 Excel = 一个流态 (class 0-4)
  每个 sheet = 一个工况点
  每列 = 一个长度为 2000 的时域样本

输出: 5_dataset.npz (或指定路径)
  含 train_X, train_y, test_X, test_y (80/20 随机分层划分)
  也支持 --no-split 模式, 存为 X, y (全部样本)

用法:
  python scripts/extract_dataset.py                          # 默认
  python scripts/extract_dataset.py --no-split --output raw_5class.npz
"""

import numpy as np
import openpyxl
from pathlib import Path
from sklearn.model_selection import train_test_split
import argparse
import sys


def extract_excel(filepath):
    """
    读取一个 Excel 文件, 返回所有有效样本 (numpy 数组, shape=[n_samples, 2000]).

    规则:
      - 遍历所有 sheet
      - 每列是一个样本, 取前 2000 行数值
      - 跳过全 NaN / 全零列 (空列或无效列)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    samples = []
    n_empty = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        max_row = ws.max_row or 2000
        max_col = ws.max_column or 0

        if max_row < 100 or max_col == 0:
            continue

        # 一次读取整个 sheet (read_only 模式下 iter_rows 最高效)
        # 构建列容器
        cols_data = [[] for _ in range(max_col)]
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 2000),
                                max_col=max_col, values_only=True):
            for ci, val in enumerate(row):
                cols_data[ci].append(val)

        for ci, col_vals in enumerate(cols_data):
            arr = np.array(col_vals, dtype=np.float32)

            # 跳过无效列
            if arr.shape[0] < 100:
                n_empty += 1
                continue
            if np.all(np.isnan(arr)):
                n_empty += 1
                continue
            # 去 NaN (极少情况)
            arr = np.nan_to_num(arr, nan=0.0)

            samples.append(arr)

    wb.close()
    if n_empty > 0:
        print(f"    跳过 {n_empty} 个无效/空列")
    return samples


def main():
    parser = argparse.ArgumentParser(description="从 Excel 提取样本到 NPZ")
    parser.add_argument('--data-dir', default='data',
                        help='Excel 文件目录 (默认: data/)')
    parser.add_argument('--output', default='5_dataset.npz',
                        help='输出 NPZ 路径 (默认: 5_dataset.npz)')
    parser.add_argument('--split', action='store_true',
                        help='做 train/test 划分 (默认: 不划分, 存 X/y)')
    parser.add_argument('--test-size', type=float, default=0.2,
                        help='测试集比例 (仅 --split 时有效, 默认: 0.2)')
    parser.add_argument('--seed', type=int, default=42,
                        help='随机种子 (默认: 42)')
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    patterns = [
        (data_dir / 'LTai1.xlsx', 0, '流态一 (空转)'),
        (data_dir / 'LTai2.xlsx', 1, '流态二 (低负荷)'),
        (data_dir / 'LTai3.xlsx', 2, '流态三 (中负荷)'),
        (data_dir / 'LTai4.xlsx', 3, '流态四 (高负荷)'),
        (data_dir / 'LTai5.xlsx', 4, '流态五 (抽水)'),
    ]

    all_X = []
    all_y = []
    class_counts = {}

    for filepath, class_id, label in patterns:
        if not filepath.exists():
            print(f"⚠ 跳过 (不存在): {filepath}")
            continue

        print(f"读取 {filepath.name} → Class {class_id} ({label}) ...")
        samples = extract_excel(filepath)

        if not samples:
            print(f"  ⚠ 未提取到有效样本")
            continue

        X_class = np.stack(samples, axis=0)   # [N, 2000]
        y_class = np.full(X_class.shape[0], class_id, dtype=np.int64)

        all_X.append(X_class)
        all_y.append(y_class)
        class_counts[class_id] = X_class.shape[0]

        print(f"  ✓ {X_class.shape[0]} 个样本 (sheets={len(samples)//293 if X_class.shape[0]%293==0 else '?'})")

    if not all_X:
        print("错误: 未提取到任何样本")
        sys.exit(1)

    X = np.concatenate(all_X, axis=0)
    y = np.concatenate(all_y, axis=0)

    print(f"\n总计: {X.shape[0]} 个样本 × {X.shape[1]} 点 = {X.shape}")
    for c, n in sorted(class_counts.items()):
        print(f"  Class {c}: {n} 样本")

    if not args.split:
        np.savez_compressed(args.output, X=X, y=y)
        print(f"\n保存到 {args.output}: X={X.shape}, y={y.shape}")
    else:
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=args.test_size, random_state=args.seed,
            stratify=y,
        )
        np.savez_compressed(
            args.output,
            train_X=X_train, train_y=y_train,
            test_X=X_test, test_y=y_test,
        )
        print(f"\n保存到 {args.output}:")
        print(f"  train_X={X_train.shape}, train_y={y_train.shape}")
        print(f"  test_X ={X_test.shape},  test_y ={y_test.shape}")

        # 验证分布一致性
        print("\n--- class 分布核对 ---")
        for c in sorted(class_counts):
            tr = (y_train == c).sum()
            te = (y_test == c).sum()
            print(f"  Class {c}: train={tr} test={te} total={tr+te}")


if __name__ == '__main__':
    main()
